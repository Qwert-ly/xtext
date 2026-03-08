from util import *
from openpyxl import load_workbook

SHEET = '上古汉语音节表.xlsx'

def char_count():
    def result_count(index, char, zhou=False, src=False):
        results = index.get(char, [])

        if src:
            return results[0][0][:-4].split('.')[-1] if results and all(result[0] == results[0][0] for result in results) and not zhou else ''
        else:
            return len(results) if results else ''
    
    W = load_workbook(filename=SHEET)
    wb = W['字典表']
    max_r = wb.max_row
    df = pd.read_parquet('data')
    idx = create_idx('ctext - 副本 - 副本', 'index.json', save=True)
    xizhou = create_idx(r'xtext\dif', 'index-xizhou.json', save=True)

    西周 = []
    次數 = []
    出處 = []
    for i in range(2, max_r + 1):
        c = wb['A' + str(i)].value
        西周.append(result_count(xizhou, c, zhou=True))
        次數.append(result_count(idx, c))
        出處.append(result_count(idx, c, src=True))
    df = pd.DataFrame({
        '總出現次數': 次數,
        '見西周': 西周,
        '少見詞出處': 出處
    })

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    for col, header in enumerate(df.columns, start=2):
        ws.cell(row=1, column=col, value=header)
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=2):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(f'_zd_语料统计.xlsx')

    noww = read_files('ctext - 副本 - 副本')[1]
    now = set(pd.read_excel(SHEET, sheet_name='字典表', header=None).iloc[:, 0].dropna())
    print(noww-now)
    print(len(noww-now))


def clean_brackets(text):
    if pd.isna(text) or text == '':
        return text
    return re.sub(r'\{[^}]*}', '', str(text))


def clean_bracket(input_file=SHEET, output_file='_xy_无括号.xlsx', sheet_name='小韻表', column='字'):
    print(f"正在读取文件 {input_file} 的工作表 '{sheet_name}'...")
    df = pd.read_excel(input_file, sheet_name=sheet_name)

    if column not in df.columns:
            print(f"错误: 工作表中不存在列 '{column}'")
            print(f"可用的列有: {list(df.columns)}")
            return False

    print(f"\n处理前 {column} 列的前5行数据:")
    print(df[column].head())
    print(f"\n正在清理 {column} 列的花括号内容...")
    df[column] = df[column].apply(clean_brackets)
    print(f"\n处理后 {column} 列的前5行数据:")
    print(df[column].head())

    df.to_excel(output_file, sheet_name=sheet_name, index=False)
    print(f"处理完成！文件已保存为 {output_file}")

def dict_(dir='ctext - 副本 - 副本'):
    """查找dir下各个txt未收于字典表的字"""
    char_list = frozenset(pd.read_excel(SHEET, sheet_name='字典表', header=None).iloc[:, 0].dropna())

    all_new_chars = set()
    with open('_new_chars.txt', 'w', encoding='utf-8') as file:
        for f in os.listdir(dir):
            if f.endswith('.txt'):
                new_c = process_text(os.path.join(dir, f), char_list)
                if len(new_c):
                    all_new_chars.update(new_c)
                    print(f'\n{f[:-4]}中{len(new_c)}个字字音不明')
                    print(''.join(sorted(new_c)))
                    file.write(f'\n{f[:-4]}中{len(new_c)}个字字音不明\n')
                    file.write(''.join(sorted(new_c)) + '\n')

        print(f'\n共{len(all_new_chars)}个')
        file.write(f'\n共{len(all_new_chars)}个\n')
        all_new_chars = ''.join(sorted(all_new_chars))
        print(all_new_chars)
        file.write(all_new_chars)


def split_syllable(ipa):
    """
    将IPA和拼音分割为声母、介音和韵母

    Args:
        ipa: IPA音标字符串

    Returns:
        tuple: (声母, 介音类别, 韵母)
    """
    # 去除末尾声调标记
    ipa_no_tone = re.sub(r'[ʔhq]$', '', ipa)
    initial = 'r̥'
    medial = ''
    final = ipa_no_tone[1:]

    # 确定介音类型
    if 'ˤr' in ipa_no_tone:
        ipa_parts = ipa_no_tone.split('ˤr', 1)
        return ipa_parts[0], '二', ipa_parts[1]
    elif 'r' in ipa_no_tone and not ipa_no_tone.startswith('r') and not ipa_no_tone.endswith('r'):
        # 如果r不在开头也不在结尾，视为介音
        ipa_parts = ipa_no_tone.split('r', 1)
        return ipa_parts[0], '三', ipa_parts[1]
    elif 'ˤ' in ipa_no_tone:
        ipa_parts = ipa_no_tone.split('ˤ', 1)
        return ipa_parts[0], '一', ipa_parts[1]

    # 特殊情况处理：找出声母和韵母的分界点
    # 使用拼音帮助判断，因为拼音的分界可能更清晰
    if ipa_no_tone.startswith('r'):
        if 'r̥' in ipa_no_tone and ipa_no_tone.startswith('r̥'):
            final = ipa_no_tone[len('r̥'):]
    elif 'C.r' in ipa_no_tone and ipa_no_tone.startswith('C.r'):
        initial = 'C.r'
        final = ipa_no_tone[3:]
    # 更复杂的情况，可能需要预定义声母表来匹配
    else:
        # 简单处理：假设第一个辅音或辅音组合是声母
        match = re.match(r'^([^aeiouəɨɯ]+)', ipa_no_tone)
        if match:
            initial = match.group(1)
            final = ipa_no_tone[len(initial):]

    return initial, medial, final


def parse():
    df = pd.read_excel(SHEET, sheet_name='小韻表', keep_default_na=False)

    # data = {
    #         'IPA': ['tsʰˤaʔ', 'pʰaŋh', 'nrap', 'ẘˤres', 'rawk', 'r̥ˤat', 'C.rəmʔ'],
    #         '拼音': ['tsh\'aq', 'phangh', 'nrap', 'wh\'res', 'rawk', 'rh\'at', 'C.rymq']
    #     }

    # df = pd.DataFrame(data)
    results = []

    for _, row in df.iterrows():
        ipa = row['IPA']

        initial, medial, final = split_syllable(ipa)

        results.append({
            'IPA': ipa,
            '拼音': row['拼音'],
            '声': initial,
            '介': medial,
            '韵': final
        })
    result_df = pd.DataFrame(results)
    print(result_df)
    result_df.to_excel('_xy_音节拆分.xlsx', index=False)


def display_results_tu(df=None, df2=None):
    res = []
    if df is not None:
        for _, r in df.iterrows():
            res.append((r.iloc[2], f'{r.iloc[7]}：{r.iloc[2]}', f'{r.iloc[8]}'))
            # print(f'{r.iloc[7]}：{r.iloc[2]}\t{r.iloc[8]}')
    # print(f'找到{len(results)}个结果\n')
    for r in res:
        sr = df2[df2.iloc[:, 3].astype(str).str.contains(r[0], na=False)]
        if sr is not None:
            for _, t in sr.iterrows():
                print(f'{r[1]} ({t.iloc[4]}小韻：{t.iloc[5]}{t.iloc[6]}切) {r[2]}')


def compare_text_vectors(text1_name, text2_name, texts, all_chars):
    if text1_name not in texts or text2_name not in texts:
        raise ValueError('文本不在texts中')

    text1, text2 = texts[text1_name], texts[text2_name]
    char_freqs = {text1_name: char_freq(text1),
                  text2_name: char_freq(text2)}

    all_chars_list = sorted(all_chars)
    vec_d = t2v(text1_name, all_chars_list, char_freqs) - t2v(text2_name, all_chars_list, char_freqs)

    char_c1 = Counter(text1)
    char_c2 = Counter(text2)
    word_d = {}
    for c, diff in zip(all_chars_list, vec_d):
        if diff != 0:
            c1 = char_c1.get(c, 0)
            c2 = char_c2.get(c, 0)
            word_d[c] = (c1-c2, diff)

    return vec_d, dict(sorted(word_d.items(), key=lambda i: i[1][1], reverse=True))


def compare_texts(text1, text2, texts, all_chars):
    _, word_d = compare_text_vectors(text1, text2, texts, all_chars)
    print(f'{text1}比{text2}更常用的字（从高到低降序）：')
    for word, diff in word_d.items():
        print(f'{word}: {diff}')


def search_worddif():
    texts, all_chars = read_files('ctext - 副本 - 副本')
    while True:
        compare_texts(input('text1:').strip(), input('text2:').strip(), texts, all_chars)


def search_tu():
    def display_tu(df=None, df2=None):
        res = []
        if df is not None:
            for _, r in df.iterrows():
                res.append((r.iloc[2], f'{r.iloc[7]}：{r.iloc[2]}', f'{r.iloc[8]}'))
                # print(f'{r.iloc[7]}：{r.iloc[2]}\t{r.iloc[8]}')
        # print(f'找到{len(results)}个结果\n')
        for r in res:
            sr = df2[df2.iloc[:, 3].astype(str).str.contains(r[0], na=False)]
            if sr is not None:
                for _, t in sr.iterrows():
                    print(f'{r[1]} ({t.iloc[4]}小韻：{t.iloc[5]}{t.iloc[6]}切) {r[2]}')
    texts_dir = 'ctext - 副本 - 副本'
    df = pd.read_parquet('tshet-uinh')
    df.columns = ['頁', '行', '音韻地位描述', '聲調', '韻目', '序数', '小韻', '字頭', '釋義']
    sr = pd.read_parquet('tshet-uinh-sr')
    sr.columns = ['小韻號', '藤田條目號', '韻目原貌', '音韻地位', '代表字', '反切上字', '反切下字', '反切上字-又', '反切下字-又', '直音',
                  '反切上字地位', '反切下字地位', '反切上字地位-又', '反切下字地位-又', '對應廣韻小韻號', '兩家差異注釋', '代表字釋義',
                  '代表字釋義-又（又 = 李永富資料之與前一列不同者）']

    while True:
        char = input("请输入要查找的字符 (打 'q' 退出): ")
        if char.lower() == 'q':
            break

        for c in char:
            display_tu(df=df[df.iloc[:, 7].astype(str).str.contains(c, na=False)], df2=sr)


def search_char():
    texts_dir = 'ctext - 副本 - 副本'
    df = pd.read_parquet('形聲考_240425')
    df.columns = ['聲首', '諧聲域', '音節類型', '二級聲符', '切拼',
                  '聲', '開合', '等', '韻', '調',
                  '字', '備註']
    idx = create_idx(texts_dir, 'index.json', save=True)

    while True:
        char = input("请输入要查找的字符 (打 'q' 退出): ")
        if char.lower() == 'q':
            break

        display_results(idx, char, texts_dir, df=df[df.iloc[:, 10].astype(str).str.contains(char, na=False)])

def expand_characters(df):
    data = []

    for _, row in df.iterrows():
        chars = row[0].split()
        try:
            parts = row[1].split()
        except:
            pass

        char = list(chars[0])  # 拆分为单个字
        pinyin = parts[0]  # 对应的拟音

        # 为每个字创建一行
        for c in char:
            data.append([c, pinyin])

    return pd.DataFrame(data, columns=['字', '音'])


def to_dict(to='_zd_output.xlsx'):
    df = pd.read_excel(SHEET, sheet_name='小韻表', keep_default_na=False)
    df['字'] = df['字'].apply(clean_brackets)
    x_df = expand_characters(df[['字', 'IPA']])

    with pd.ExcelWriter(to, engine='openpyxl') as w:
        x_df.to_excel(w, index=False)

def main_panel():
    while True:
        print(" [1] 语料字频统计")
        print("     - 统计字的词频")
        print(" [2] 清理Excel花括号")
        print("     - 去除指定工作表及列中的 {xxx} 内容")
        print(" [3] 查找未收录生僻字")
        print("     - 检查未在字典表中的字")
        print(" [4] 音节拆分: 声/介/韵")
        print("     - 读取小韻表，拆分IPA和拼音")
        print(" [5] 切韵音查询")
        print("     - 检索切韵(藤田复原)")
        print(" [6] 文本用字差异比较")
        print("     - 对比文本的字频向量差异")
        print(" [7] 形声考+语料查询")
        print("     - 检索汉字的形声考地位及语料出处")
        print(" [8] 小韵表转字典表")

        choice = input("请输入功能编号: ").strip()

        if choice == '1':
            print("\n语料字频统计...")
            char_count()
            print("已保存至_zd_语料统计.xlsx")

        elif choice == '2':
            print("\n清理Excel花括号...")
            in_file = input("  输入文件名 (默认 上古汉语音节表.xlsx): ").strip() or SHEET
            out_file = input("  输出文件名 (默认 _xy_无括号.xlsx): ").strip() or '_xy_无括号.xlsx'
            sheet = input("  工作表名 (直接回车默认 '小韻表'): ").strip() or '小韻表'
            col = input("  要清理的列字母 (直接回车默认 '字'): ").strip() or '字'
            clean_bracket(in_file, out_file, sheet, col)

        elif choice == '3':
            print("\n查找未收录字...")
            target_dir = input(r"  txt所在目录 (直接回车默认 'ctext - 副本 - 副本'): ").strip()
            if target_dir:
                dict_(dir=target_dir)
            else:
                dict_()
            print("结果已保存至 _new_chars.txt")

        elif choice == '4':
            print("\n音节拆分...")
            parse()
            print("结果已保存至 _xy_音节拆分.xlsx")

        elif choice == '5':
            print("\n切韵查询 (打 'q' 退出)")
            search_tu()

        elif choice == '6':
            print("\n文本用字差异比较 (按 Ctrl+C 退出)")
            search_worddif()

        elif choice == '7':
            print("\n语料+形声考查询 (打 'q' 退出)")
            search_char()

        elif choice == '8':
            print("\n小韵表转字典表")
            target_dir = input(r"  输出文件 (直接回车默认 '_zd_output.xlsx'): ").strip()
            if target_dir:
                to_dict(target_dir)
            else:
                to_dict()
            print(f"结果已保存至{target_dir}")

        else:
            print("\n[错误] 无效的输入，请重新输入 0 到 8 之间的数字！")


if __name__ == "__main__":
    main_panel()
